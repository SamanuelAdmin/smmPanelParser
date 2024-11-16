import openpyxl
import os

class ExcelController:
    @staticmethod
    def load_excel_file(path: str):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        return sheet.iter_rows()

    @staticmethod
    def export_excel_file(query: list, path: str, filename: str):
        wb = openpyxl.Workbook()
        sheet = wb.active

        num_row = 0
        for row in query:
            num_row += 1
            for col in range(1, 2 + 1):
                value = row[col]
                sheet.cell(num_row, col, value)
        wb.save(filename=os.path.join(path, f'{filename if filename else "export"}.xlsx'))